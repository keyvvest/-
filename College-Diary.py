import sys
import sqlite3
import bcrypt
from PyQt6.QtWidgets import *
from PyQt6.QtCore import Qt, QSize, QDate
from PyQt6.QtCore import Qt, QSize, pyqtSignal, QThread
from PyQt6.QtGui import QFont
from PyQt6.QtGui import QRegularExpressionValidator
from PyQt6.QtCore import QRegularExpression
from functools import partial
import datetime
from datetime import datetime
import secrets
import os
import re
import openpyxl
from openpyxl.styles import Font, Alignment


db_path = "../../Downloads/College.db"

# Описание новых цветов для спокойного дизайна
light_gray = "#F4F4F9"
dark_gray = "#A3A3A3"
light_blue = "#B3C7E6"
dark_blue = "#2F4F7F"
light_green = "#A7D8A0"
dark_green = "#2C6A4F"

# Новый стиль для приложения
app_style = f"""
QWidget {{
    background-color: {light_gray};
    font-family: 'Arial', sans-serif;
    font-size: 12pt;
    color: black;  /* Устанавливаем цвет текста на черный */
}}

QPushButton {{
    background-color: {light_blue};
    border: 2px solid {dark_blue};
    border-radius: 8px;
    padding: 8px 16px;
    min-height: 40px;
    color: white;
    font-weight: bold;
}}

QPushButton:hover {{
    background-color: {dark_blue};
    border: 2px solid {light_blue};
}}

QPushButton:pressed {{
    background-color: {dark_blue};
}}

QLabel {{
    font-size: 14pt;
    color: {dark_gray};
    font-weight: bold;
}}

QLineEdit {{
    border: 1px solid {dark_gray};
    border-radius: 5px;
    padding: 8px;
    min-height: 35px;
    background-color: white;
    color: black;  /* Устанавливаем цвет текста на черный */
}}

QComboBox {{
    border: 1px solid {dark_gray};
    border-radius: 5px;
    padding: 8px;
    background-color: white;
    min-width: 200px;  /* Увеличим ширину для удобства */
    color: black;  /* Устанавливаем цвет текста на черный */
}}

QComboBox QAbstractItemView {{
    border: 1px solid {dark_gray};
    background-color: white;
    color: black;  /* Устанавливаем цвет текста на черный */
    selection-background-color: {light_blue};  /* Цвет фона выделенного элемента */
    selection-color: white;  /* Цвет текста выделенного элемента */
}}

QTableWidget {{
    border: 1px solid {dark_gray};
    background-color: white;
    gridline-color: {dark_gray};
    color: black;  /* Устанавливаем цвет текста на черный */
}}

QTableWidgetItem {{
    padding: 10px;
    color: black;  /* Устанавливаем цвет текста на черный */
}}

QTabWidget::pane {{
    border: 1px solid {dark_gray};
    background-color: {light_gray};
    color: black;  /* Устанавливаем цвет текста на черный */
}}

QTabBar::tab {{
    background-color: {light_gray};
    border: 1px solid {dark_gray};
    padding: 8px 16px;
    min-width: 100px;
    color: {dark_gray};
}}

QTabBar::tab:selected {{
    background-color: {light_blue};
    color: white;
}}

QTabBar::tab:hover {{
    background-color: {dark_blue};
    color: white;
}}

QMessageBox {{
    background-color: {light_gray};
    color: {dark_gray};
}}

QMessageBox QLabel {{
    color: {dark_gray};
}}

QMessageBox QPushButton {{
    background-color: {light_blue};
    color: white;
    border: 1px solid {dark_blue};
    border-radius: 5px;
    padding: 5px 10px;
}}

QMessageBox QPushButton:hover {{
    background-color: {dark_blue};
}}

QListWidget {{
    border: 1px solid {dark_gray};
    background-color: white;
    color: black;  /* Устанавливаем цвет текста на черный */
}}

QListWidget::item {{
    padding: 5px;
    color: black;  /* Устанавливаем цвет текста на черный */
}}

QListWidget::item:selected {{
    background-color: {light_blue};
    color: white;  /* Цвет текста выделенного элемента */
}}

QListWidget::item:hover {{
    background-color: {light_blue};  /* Цвет фона при наведении */
    color: white;  /* Цвет текста при наведении */
}}
"""

import logging
logging.basicConfig(level=logging.DEBUG)


class StudentWindow(QWidget):
    def __init__(self, student_id, db_path, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Окно студента")
        self.db_path = db_path
        self.student_id = student_id
        self.setMinimumSize(800, 600)

        # Fetch student information and disciplines in the background
        self.student_info = self.get_student_info()
        self.disciplines = self.get_student_disciplines()

        # Create labels for student information
        self.student_label = QLabel(f"Студент: {self.student_info['fio']}")
        self.group_label = QLabel(f"Группа: {self.student_info['group_name']}")
        self.specialty_label = QLabel(f"Специальность: {self.student_info['speciality_name']}")

        # Remove the discipline combo box (as requested)
        # No need to change discipline by combo anymore

        # Create journal table
        self.journal_table = QTableWidget()
        self.journal_table.setColumnCount(3)  # Three columns: Discipline, Date, Grade
        self.journal_table.setHorizontalHeaderLabels(["Дисциплина", "Дата", "Оценка"])

        # Layout
        layout = QVBoxLayout()
        layout.addWidget(self.student_label)
        layout.addWidget(self.group_label)
        layout.addWidget(self.specialty_label)
        layout.addWidget(self.journal_table)
        self.setLayout(layout)

        # Populate journal with grades and disciplines
        self.populate_journal()

    def get_student_info(self):
        try:
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()
            cursor.execute("""
                SELECT s.last_name, s.first_name, s.patronymic, g.group_name, g.speciality_name
                FROM Student s
                JOIN Group_ g ON s.group_id = g.group_id
                WHERE s.user_id = ?
            """, (self.student_id,))
            result = cursor.fetchone()
            conn.close()
            if result:
                last_name, first_name, patronymic, group_name, speciality_name = result
                fio = f"{last_name} {first_name} {patronymic}"
                return {
                    'fio': fio,
                    'group_name': group_name,
                    'speciality_name': speciality_name
                }
            else:
                logging.warning(f"Студент с ID {self.student_id} не найден")
                return {
                    'fio': "Не найден",
                    'group_name': "Не найдена",
                    'speciality_name': "Не найдена"
                }
        except sqlite3.Error as e:
            logging.error(f"Ошибка базы данных при получении информации о студенте: {e}")
            return {
                'fio': "Ошибка",
                'group_name': "Ошибка",
                'speciality_name': "Ошибка"
            }

    def get_student_disciplines(self):
        try:
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()
            cursor.execute("""
                SELECT d.discipline_id, d.discipline_name
                FROM Discipline d
                JOIN Group_ g ON d.group_id = g.group_id
                WHERE g.group_id = (SELECT group_id FROM Student WHERE user_id = ?)
            """, (self.student_id,))
            results = cursor.fetchall()
            conn.close()
            return [{'id': result[0], 'name': result[1]} for result in results]
        except sqlite3.Error as e:
            logging.error(f"Ошибка базы данных при получении дисциплин студента: {e}")
            return []

    def get_grades_for_discipline(self, discipline_id):
        try:
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()
            cursor.execute("""
                SELECT grade_date, grade_value, grade_type
                FROM Grade
                WHERE student_id = ? AND discipline_id = ?
            """, (self.student_id, discipline_id))
            grades = cursor.fetchall()
            conn.close()
            return grades
        except sqlite3.Error as e:
            logging.error(f"Ошибка базы данных при получении оценок: {e}")
            return []

    def populate_journal(self):
        # Сначала очищаем таблицу
        self.journal_table.setRowCount(0)

        # Получаем уникальные даты всех дисциплин
        all_dates = set()
        for discipline in self.disciplines:
            grades = self.get_grades_for_discipline(discipline['id'])
            for grade in grades:
                date = grade[0]
                if isinstance(date, datetime):
                    all_dates.add(date.strftime("%Y-%m-%d"))
                else:
                    all_dates.add(str(date))

        # Сортируем даты
        sorted_dates = sorted(all_dates)

        # Устанавливаем количество столбцов: 1 для дисциплины, затем для дат, и 1 для средней оценки
        self.journal_table.setColumnCount(len(sorted_dates) + 2)  # +2: 1 для дисциплины и 1 для средней оценки
        self.journal_table.setHorizontalHeaderLabels(['Дисциплина'] + sorted_dates + ['Средняя оценка'])

        # Заполняем таблицу
        for row, discipline in enumerate(self.disciplines):
            grades = self.get_grades_for_discipline(discipline['id'])

            # Добавляем новую строку для каждой дисциплины
            self.journal_table.insertRow(row)

            # Заполняем столбец с названием дисциплины
            discipline_item = QTableWidgetItem(discipline['name'])
            discipline_item.setFlags(
                discipline_item.flags() & ~Qt.ItemFlag.ItemIsEditable)  # Делаем название дисциплины не редактируемым
            self.journal_table.setItem(row, 0, discipline_item)

            # Заполняем оценки по датам
            for col, date in enumerate(sorted_dates):
                # Получаем оценку для этой дисциплины и даты
                grade_value = None
                grade_type = None
                for grade in grades:
                    grade_date, grade_value_in_db, grade_type_in_db = grade  # Теперь возвращаем и тип оценки
                    if isinstance(grade_date, datetime):
                        grade_date = grade_date.strftime("%Y-%m-%d")

                    # Если дата совпадает, то берем оценку и тип
                    if grade_date == date:
                        grade_value = grade_value_in_db
                        grade_type = grade_type_in_db
                        break

                # Если оценка найдена, вставляем её, иначе оставляем ячейку пустой
                if grade_value:
                    grade_text = str(grade_value)
                    grade_item = QTableWidgetItem(grade_text)

                    # Устанавливаем подсказку (tooltip) с информацией о типе оценки
                    if grade_type:
                        grade_item.setToolTip(f"{grade_type}")
                else:
                    grade_item = QTableWidgetItem()  # Оставляем ячейку пустой

                # Делаем ячейку с оценкой не редактируемой
                grade_item.setFlags(grade_item.flags() & ~Qt.ItemFlag.ItemIsEditable)

                self.journal_table.setItem(row, col + 1, grade_item)  # +1 для смещения в столбцы с датами

            # Расчет средней оценки для студента по дисциплине
            average_grade = self.calculate_average_grade(grades)

            # Столбец со средней оценкой
            average_grade_item = QTableWidgetItem(f"{average_grade:.2f}")

            # Запрещаем редактирование ячейки средней оценки
            average_grade_item.setFlags(average_grade_item.flags() & ~Qt.ItemFlag.ItemIsEditable)

            self.journal_table.setItem(row, len(sorted_dates) + 1, average_grade_item)

        # Устанавливаем ширину столбцов
        self.journal_table.resizeColumnsToContents()

        # Устанавливаем фиксированную ширину для столбцов с датами (можно настроить под нужды)
        for col in range(1, len(sorted_dates) + 1):
            self.journal_table.setColumnWidth(col, 100)  # 100 пикселей для столбцов с датами

        # Устанавливаем фиксированную ширину для столбца с дисциплиной
        self.journal_table.setColumnWidth(0, 200)  # 200 пикселей для столбца с дисциплиной

        # Устанавливаем большую ширину для столбца с "Средняя оценка"
        self.journal_table.setColumnWidth(len(sorted_dates) + 1, 150)  # 150 пикселей для столбца "Средняя оценка"

        # Если по какой-то причине добавился лишний столбец (например, столбец с ФИО), убираем его (предпоследний столбец)
        if self.journal_table.columnCount() > 8:
            self.journal_table.removeColumn(self.journal_table.columnCount() - 2)  # Удаляем предпоследний столбец

    def calculate_average_grade(self, grades):
        """
        Рассчитывает среднюю оценку с учетом удвоенных оценок за контрольные работы.
        """
        total_grades = []
        for grade_date, grade_value, grade_type in grades:
            if grade_value is not None:
                total_grades.append(grade_value)
                # Если это контрольная работа, то добавляем оценку еще раз
                if "Контрольная работа" in grade_type:  # Предполагаем, что "Контрольная работа" в типе оценки
                    total_grades.append(grade_value)

        if total_grades:
            return sum(total_grades) / len(total_grades)
        else:
            return 0







class AdminWindow(QWidget):
    def __init__(self, db_path, parent=None):
        super().__init__(parent)
        self.db_path = db_path
        self.setWindowTitle("Окно администратора")
        self.setMinimumSize(800, 600)

        # Cache for students and groups
        self.students_cache = []
        self.groups_cache = []

        # Create tabs for different admin functionalities
        self.username_edit = QLineEdit()
        self.password_edit = QLineEdit()
        self.first_name_edit = QLineEdit()
        self.last_name_edit = QLineEdit()
        self.patronymic_edit = QLineEdit()
        self.group_select_combo = QComboBox()
        self.teacher_select_combo = QComboBox()
        self.student_select_combo = QComboBox()
        self.edit_teacher_select_combo = QComboBox()
        self.edit_student_select_combo = QComboBox()
        self.move_student_select_combo = QComboBox()
        self.group_select_combo = QComboBox()
        self.edit_group_select_combo = QComboBox()
        self.tab_widget = QTabWidget()
        self.add_student_tab = self.create_add_student_tab()
        self.add_teacher_tab = self.create_add_teacher_tab()
        self.delete_student_tab = self.create_delete_student_tab()
        self.delete_teacher_tab = self.create_delete_teacher_tab()
        self.add_group_tab = self.create_add_group_tab()
        self.delete_group_tab = self.create_delete_group_tab()
        self.move_student_tab = self.create_move_student_tab()
        self.edit_student_tab = self.create_edit_student_tab()  # Новая вкладка для редактирования студентов
        self.edit_teacher_tab = self.create_edit_teacher_tab()  # Новая вкладка для редактирования преподавателей
        self.tab_widget.addTab(self.add_student_tab, "Добавить студента")
        self.tab_widget.addTab(self.add_teacher_tab, "Добавить преподавателя")
        self.tab_widget.addTab(self.delete_student_tab, "Удалить студента")
        self.tab_widget.addTab(self.delete_teacher_tab, "Удалить преподавателя")
        self.tab_widget.addTab(self.add_group_tab, "Добавить группу")
        self.tab_widget.addTab(self.delete_group_tab, "Удалить группу")
        self.tab_widget.addTab(self.move_student_tab, "Переместить студента")
        self.tab_widget.addTab(self.edit_student_tab, "Редактировать студента")  # Добавляем новую вкладку
        self.tab_widget.addTab(self.edit_teacher_tab, "Редактировать преподавателя")  # Добавляем новую вкладку

        layout = QVBoxLayout()
        layout.addWidget(self.tab_widget)
        self.setLayout(layout)

        # Load initial data
        self.load_students(self.student_select_combo)
        self.load_groups(self.group_select_combo)

    def show_question_message(self, title, message):
        messageBox = QMessageBox(self)
        messageBox.setIcon(QMessageBox.Icon.Question)
        messageBox.setText(message)
        messageBox.setWindowTitle(title)

        yesButton = QPushButton("Да")
        noButton = QPushButton("Нет")

        messageBox.addButton(yesButton, QMessageBox.ButtonRole.YesRole)
        messageBox.addButton(noButton, QMessageBox.ButtonRole.NoRole)

        reply = messageBox.exec()

        if messageBox.clickedButton() == yesButton:
            return True
        else:
            return False

    def show_critical_message(self, title, message):
        messageBox = QMessageBox(self)
        messageBox.setIcon(QMessageBox.Icon.Critical)
        messageBox.setText(message)
        messageBox.setWindowTitle(title)

        okButton = QPushButton("ОК")
        messageBox.addButton(okButton, QMessageBox.ButtonRole.AcceptRole)

        messageBox.exec()

    def show_warning_message(self, title, message):
        messageBox = QMessageBox(self)
        messageBox.setIcon(QMessageBox.Icon.Warning)
        messageBox.setText(message)
        messageBox.setWindowTitle(title)

        okButton = QPushButton("ОК")
        messageBox.addButton(okButton, QMessageBox.ButtonRole.AcceptRole)

        messageBox.exec()

    def show_info_message(self, title, message):
        messageBox = QMessageBox(self)
        messageBox.setIcon(QMessageBox.Icon.Information)
        messageBox.setText(message)
        messageBox.setWindowTitle(title)

        okButton = QPushButton("ОК")
        messageBox.addButton(okButton, QMessageBox.ButtonRole.AcceptRole)

        messageBox.exec()

    def update_lists(self):
        self.load_students(self.edit_student_select_combo)  # Обновляем список студентов для редактирования
        self.load_students(self.student_select_combo)  # Обновляем список студентов для удаления
        self.load_students_for_move(self.move_student_list_widget)  # Обновляем список студентов для перемещения
        self.load_teachers(self.edit_teacher_select_combo)  # Обновляем список преподавателей для редактирования
        self.load_teachers(self.teacher_select_combo)  # Обновляем список преподавателей для удаления
        self.load_groups(self.group_select_combo)  # Обновляем список групп
        self.load_groups_for_move(self.move_group_list_widget)  # Обновляем список групп для перемещения студентов
        self.load_students_to_list(self.student_list_widget)  # Обновляем список студентов
        self.load_teachers_to_list(self.teacher_list_widget)  # Обновляем список преподавателей
        self.load_groups_to_list(self.group_list_widget)  # Обновляем список групп

    def generate_password(self, password_input):
        password = secrets.token_urlsafe(9)
        password_input.setText(password)

    def create_add_student_tab(self):
        tab = QWidget()
        layout = QVBoxLayout()

        group_box = QGroupBox("Добавить студента")
        group_box.setStyleSheet("QGroupBox { margin-top: 1ex; }"
                                "QGroupBox::title { subcontrol-origin: margin; left: 10px; }")

        form_layout = QFormLayout()
        form_layout.setLabelAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignTop)
        form_layout.setContentsMargins(20, 10, 20, 10)  # Adjust margins
        form_layout.setVerticalSpacing(20)  # Increase vertical spacing

        self.student_fname_input = QLineEdit()
        form_layout.addRow(QLabel("Имя:"), self.student_fname_input)

        self.student_lname_input = QLineEdit()
        form_layout.addRow(QLabel("Фамилия:"), self.student_lname_input)

        self.student_patronymic_input = QLineEdit()
        form_layout.addRow(QLabel("Отчество:"), self.student_patronymic_input)

        self.student_group_input = QLineEdit()
        form_layout.addRow(QLabel("Номер группы:"), self.student_group_input)

        self.student_username_input = QLineEdit()
        form_layout.addRow(QLabel("Логин:"), self.student_username_input)

        self.student_password_input = QLineEdit()
        self.student_password_input.setEchoMode(QLineEdit.EchoMode.Normal)  # Display plain text
        form_layout.addRow(QLabel("Пароль:"), self.student_password_input)

        generate_password_button = QPushButton("Сгенерировать пароль")
        generate_password_button.clicked.connect(lambda: self.generate_password(self.student_password_input))
        form_layout.addRow(generate_password_button)

        button_layout = QHBoxLayout()
        add_student_button = QPushButton("Добавить студента")
        add_student_button.clicked.connect(self.add_student)
        button_layout.addWidget(add_student_button)

        clear_fields_button = QPushButton("Очистить поля")
        clear_fields_button.clicked.connect(self.clear_student_fields)
        button_layout.addWidget(clear_fields_button)

        form_layout.addRow(button_layout)

        group_box_layout = QVBoxLayout()
        group_box_layout.addLayout(form_layout)
        group_box.setLayout(group_box_layout)

        layout.addWidget(group_box)
        tab.setLayout(layout)
        return tab

    def clear_student_fields(self):
        self.student_fname_input.clear()
        self.student_lname_input.clear()
        self.student_patronymic_input.clear()
        self.student_group_input.clear()
        self.student_username_input.clear()
        self.student_password_input.clear()

    def add_student(self):
        fname = self.student_fname_input.text()
        lname = self.student_lname_input.text()
        patronymic = self.student_patronymic_input.text()
        group_name = self.student_group_input.text()
        login = self.student_username_input.text()
        password = self.student_password_input.text()

        # Проверка на пустые поля
        if not fname or not lname or not patronymic or not group_name or not login or not password:
            QMessageBox.warning(self, "Ошибка", "Все поля должны быть заполнены.")
            return

        # Проверка на текстовые поля
        if not fname.isalpha() or not lname.isalpha() or not patronymic.isalpha():
            QMessageBox.warning(self, "Ошибка", "Поля 'Имя', 'Фамилия' и 'Отчество' должны содержать только текст.")
            return

        # Проверка на уникальность логина
        try:
            with sqlite3.connect(self.db_path) as conn:
                cursor = conn.cursor()
                cursor.execute("SELECT COUNT(*) FROM Users WHERE username = ?", (login,))
                if cursor.fetchone()[0] > 0:
                    QMessageBox.warning(self, "Ошибка", "Пользователь с таким логином уже существует.")
                    return

                # Fetch group_id based on group_name from Group_ table
                cursor.execute("SELECT group_id FROM Group_ WHERE group_name = ?", (group_name,))
                result = cursor.fetchone()
                if result:
                    group_id = result[0]
                else:
                    QMessageBox.warning(self, "Ошибка", "Группа не найдена.")
                    return

                # Insert into Users table
                hashed_password = bcrypt.hashpw(password.encode(), bcrypt.gensalt())
                cursor.execute("INSERT INTO Users (username, password, role) VALUES (?, ?, ?)",
                               (login, hashed_password, 'student'))
                user_id = cursor.lastrowid

                # Insert into Student table
                cursor.execute(
                    "INSERT INTO Student (user_id, group_id, first_name, last_name, patronymic) VALUES (?, ?, ?, ?, ?)",
                    (user_id, group_id, fname, lname, patronymic))
                conn.commit()
                QMessageBox.information(self, "Успех", "Студент добавлен успешно.")
                self.update_lists()

                # Очистка полей
                self.student_fname_input.clear()
                self.student_lname_input.clear()
                self.student_patronymic_input.clear()
                self.student_group_input.clear()
                self.student_username_input.clear()
                self.student_password_input.clear()
        except sqlite3.Error as e:
            QMessageBox.critical(self, "Ошибка", f"Ошибка базы данных: {e}")

    def create_add_teacher_tab(self):
        tab = QWidget()
        layout = QVBoxLayout()

        group_box = QGroupBox("Добавить преподавателя")
        group_box.setStyleSheet("QGroupBox { margin-top: 1ex; }"
                                "QGroupBox::title { subcontrol-origin: margin; left: 10px; }")

        form_layout = QFormLayout()
        form_layout.setLabelAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignTop)
        form_layout.setContentsMargins(20, 10, 20, 10)  # Adjust margins
        form_layout.setVerticalSpacing(20)  # Increase vertical spacing

        self.teacher_fname_input = QLineEdit()
        form_layout.addRow(QLabel("Имя:"), self.teacher_fname_input)

        self.teacher_lname_input = QLineEdit()
        form_layout.addRow(QLabel("Фамилия:"), self.teacher_lname_input)

        self.teacher_patronymic_input = QLineEdit()
        form_layout.addRow(QLabel("Отчество:"), self.teacher_patronymic_input)

        self.teacher_username_input = QLineEdit()
        form_layout.addRow(QLabel("Логин:"), self.teacher_username_input)

        self.teacher_password_input = QLineEdit()
        self.teacher_password_input.setEchoMode(QLineEdit.EchoMode.Normal)  # Display plain text
        form_layout.addRow(QLabel("Пароль:"), self.teacher_password_input)

        generate_password_button = QPushButton("Сгенерировать пароль")
        generate_password_button.clicked.connect(lambda: self.generate_password(self.teacher_password_input))
        form_layout.addRow(generate_password_button)

        button_layout = QHBoxLayout()
        add_teacher_button = QPushButton("Добавить преподавателя")
        add_teacher_button.clicked.connect(self.add_teacher)
        button_layout.addWidget(add_teacher_button)

        clear_fields_button = QPushButton("Очистить поля")
        clear_fields_button.clicked.connect(self.clear_teacher_fields)
        button_layout.addWidget(clear_fields_button)

        form_layout.addRow(button_layout)

        group_box_layout = QVBoxLayout()
        group_box_layout.addLayout(form_layout)
        group_box.setLayout(group_box_layout)

        layout.addWidget(group_box)
        tab.setLayout(layout)
        return tab

    def clear_teacher_fields(self):
        self.teacher_fname_input.clear()
        self.teacher_lname_input.clear()
        self.teacher_patronymic_input.clear()
        self.teacher_username_input.clear()
        self.teacher_password_input.clear()

    def add_teacher(self):
        fname = self.teacher_fname_input.text()
        lname = self.teacher_lname_input.text()
        patronymic = self.teacher_patronymic_input.text()
        login = self.teacher_username_input.text()
        password = self.teacher_password_input.text().encode()

        # Проверка на пустые поля
        if not fname or not lname or not patronymic or not login or not password:
            QMessageBox.warning(self, "Ошибка", "Все поля должны быть заполнены.")
            return

        # Проверка на текстовые поля
        if not fname.isalpha() or not lname.isalpha() or not patronymic.isalpha():
            QMessageBox.warning(self, "Ошибка", "Поля 'Имя', 'Фамилия' и 'Отчество' должны содержать только текст.")
            return

        # Проверка на уникальность логина
        try:
            with sqlite3.connect(self.db_path) as conn:
                cursor = conn.cursor()
                cursor.execute("SELECT COUNT(*) FROM Users WHERE username = ?", (login,))
                if cursor.fetchone()[0] > 0:
                    QMessageBox.warning(self, "Ошибка", "Пользователь с таким логином уже существует.")
                    return

                # Insert into Users table
                hashed_password = bcrypt.hashpw(password, bcrypt.gensalt())
                cursor.execute("INSERT INTO Users (username, password, role) VALUES (?, ?, ?)",
                               (login, hashed_password, 'teacher'))
                user_id = cursor.lastrowid

                # Insert into Teacher table
                cursor.execute("INSERT INTO Teacher (user_id, first_name, last_name, patronymic) VALUES (?, ?, ?, ?)",
                               (user_id, fname, lname, patronymic))
                conn.commit()
                QMessageBox.information(self, "Успех", "Преподаватель добавлен успешно.")
                self.update_lists()  # Обновляем список преподавателей для удаления

                # Очистка полей
                self.teacher_fname_input.clear()
                self.teacher_lname_input.clear()
                self.teacher_patronymic_input.clear()
                self.teacher_username_input.clear()
                self.teacher_password_input.clear()
        except sqlite3.Error as e:
            QMessageBox.critical(self, "Ошибка", f"Ошибка базы данных: {e}")

    def create_delete_student_tab(self):
        tab = QWidget()
        layout = QVBoxLayout()

        group_box = QGroupBox("Удалить студента")
        group_box.setStyleSheet("QGroupBox { margin-top: 1ex; }"
                                "QGroupBox::title { subcontrol-origin: margin; left: 10px; }")

        form_layout = QFormLayout()
        form_layout.setLabelAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignTop)
        form_layout.setContentsMargins(20, 10, 20, 10)  # Adjust margins
        form_layout.setVerticalSpacing(20)  # Increase vertical spacing

        self.student_list_widget = QListWidget()
        self.load_students_to_list(self.student_list_widget)
        self.student_list_widget.setToolTip("Выберите студента для удаления")
        form_layout.addRow(QLabel("Выберите студента:"), self.student_list_widget)

        delete_student_button = QPushButton("Удалить студента")
        delete_student_button.clicked.connect(self.delete_student_from_list)
        delete_student_button.setToolTip("Нажмите, чтобы удалить выбранного студента")
        form_layout.addRow(delete_student_button)

        group_box_layout = QVBoxLayout()
        group_box_layout.addLayout(form_layout)
        group_box.setLayout(group_box_layout)

        layout.addWidget(group_box)
        tab.setLayout(layout)
        return tab

    def load_students_to_list(self, list_widget):
        list_widget.clear()
        try:
            with sqlite3.connect(self.db_path) as conn:
                cursor = conn.cursor()
                cursor.execute("""
                    SELECT Users.username, Student.first_name, Student.last_name, Student.user_id
                    FROM Users
                    JOIN Student ON Users.id = Student.user_id
                """)
                self.students_cache = cursor.fetchall()
                for student in self.students_cache:
                    display = f"{student[0]} - {student[1]} {student[2]}"
                    list_widget.addItem(display)
                    list_widget.item(list_widget.count() - 1).setData(Qt.ItemDataRole.UserRole, student[3])
        except sqlite3.Error as e:
            QMessageBox.critical(self, "Ошибка", f"Ошибка базы данных: {e}")

    def delete_student_from_list(self):
        selected_item = self.student_list_widget.currentItem()
        if not selected_item:
            QMessageBox.warning(self, "Ошибка", "Нет выбранного студента.")
            return
        user_id = selected_item.data(Qt.ItemDataRole.UserRole)
        self.delete_student(user_id)
        self.update_lists()

    def load_students(self, combo_box):
        combo_box.clear()
        try:
            with sqlite3.connect(self.db_path) as conn:
                cursor = conn.cursor()
                cursor.execute("""
                    SELECT Users.username, Student.first_name, Student.last_name, Student.user_id
                    FROM Users
                    JOIN Student ON Users.id = Student.user_id
                """)
                self.students_cache = cursor.fetchall()
                for student in self.students_cache:
                    display = f"{student[0]} - {student[1]} {student[2]}"
                    combo_box.addItem(display, student[3])
        except sqlite3.Error as e:
            QMessageBox.critical(self, "Ошибка", f"Ошибка базы данных: {e}")

    def create_delete_teacher_tab(self):
        tab = QWidget()
        layout = QVBoxLayout()

        group_box = QGroupBox("Удалить преподавателя")
        group_box.setStyleSheet("QGroupBox { margin-top: 1ex; }"
                                "QGroupBox::title { subcontrol-origin: margin; left: 10px; }")

        form_layout = QFormLayout()
        form_layout.setLabelAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignTop)
        form_layout.setContentsMargins(20, 10, 20, 10)  # Adjust margins
        form_layout.setVerticalSpacing(20)  # Increase vertical spacing

        self.teacher_list_widget = QListWidget()
        self.load_teachers_to_list(self.teacher_list_widget)
        self.teacher_list_widget.setToolTip("Выберите преподавателя для удаления")
        form_layout.addRow(QLabel("Выберите преподавателя:"), self.teacher_list_widget)

        delete_teacher_button = QPushButton("Удалить преподавателя")
        delete_teacher_button.clicked.connect(self.delete_teacher_from_list)
        delete_teacher_button.setToolTip("Нажмите, чтобы удалить выбранного преподавателя")
        form_layout.addRow(delete_teacher_button)

        group_box_layout = QVBoxLayout()
        group_box_layout.addLayout(form_layout)
        group_box.setLayout(group_box_layout)

        layout.addWidget(group_box)
        tab.setLayout(layout)
        return tab

    def load_teachers_to_list(self, list_widget):
        list_widget.clear()
        try:
            with sqlite3.connect(self.db_path) as conn:
                cursor = conn.cursor()
                cursor.execute("""
                    SELECT Users.username, Teacher.first_name, Teacher.last_name, Teacher.user_id
                    FROM Users
                    JOIN Teacher ON Users.id = Teacher.user_id
                """)
                teachers = cursor.fetchall()
                for teacher in teachers:
                    display = f"{teacher[0]} - {teacher[1]} {teacher[2]}"
                    list_widget.addItem(display)
                    list_widget.item(list_widget.count() - 1).setData(Qt.ItemDataRole.UserRole, teacher[3])
        except sqlite3.Error as e:
            QMessageBox.critical(self, "Ошибка", f"Ошибка базы данных: {e}")

    def delete_teacher_from_list(self):
        selected_item = self.teacher_list_widget.currentItem()
        if not selected_item:
            QMessageBox.warning(self, "Ошибка", "Нет выбранного преподавателя.")
            return
        user_id = selected_item.data(Qt.ItemDataRole.UserRole)
        self.delete_teacher(user_id)
        self.update_lists()

    def load_teachers(self, combo_box):
        combo_box.clear()
        try:
            with sqlite3.connect(self.db_path) as conn:
                cursor = conn.cursor()
                cursor.execute("""
                    SELECT Users.username, Teacher.first_name, Teacher.last_name, Teacher.user_id
                    FROM Users
                    JOIN Teacher ON Users.id = Teacher.user_id
                """)
                teachers = cursor.fetchall()
                for teacher in teachers:
                    display = f"{teacher[0]} - {teacher[1]} {teacher[2]}"
                    combo_box.addItem(display, teacher[3])
        except sqlite3.Error as e:
            QMessageBox.critical(self, "Ошибка", f"Ошибка базы данных: {e}")

    def delete_student(self, user_id):
        if not self.show_question_message("Подтверждение", "Вы уверены, что хотите удалить этого студента?"):
            return

        try:
            with sqlite3.connect(self.db_path) as conn:
                conn.execute("PRAGMA foreign_keys = ON")
                cursor = conn.cursor()
                # Start transaction
                cursor.execute("BEGIN TRANSACTION")
                # Delete related records from Grade
                cursor.execute("DELETE FROM Grade WHERE student_id = ?", (user_id,))
                print("Grades deleted.")
                # Delete related records from Attendance
                cursor.execute("DELETE FROM Attendance WHERE student_id = ?", (user_id,))
                print("Attendance deleted.")
                # Delete from Student table
                cursor.execute("DELETE FROM Student WHERE user_id = ?", (user_id,))
                print("Student deleted.")
                # Delete from Users table
                cursor.execute("DELETE FROM Users WHERE id = ?", (user_id,))
                print("User deleted.")
                conn.commit()
                QMessageBox.information(self, "Успех", "Студент удален успешно.")
                self.update_lists()  # Обновляем список студентов
        except sqlite3.IntegrityError as e:
            conn.rollback()
            QMessageBox.critical(self, "Ошибка", f"Не удалось удалить студента: {e}")
        except sqlite3.Error as e:
            conn.rollback()
            QMessageBox.critical(self, "Ошибка", f"Ошибка базы данных: {e}")
        except Exception as e:
            conn.rollback()
            QMessageBox.critical(self, "Ошибка", f"Произошла ошибка: {e}")

    def delete_teacher(self, user_id):
        if not self.show_question_message("Подтверждение", "Вы уверены, что хотите удалить этого преподавателя?"):
            return

        try:
            with sqlite3.connect(self.db_path) as conn:
                conn.execute("PRAGMA foreign_keys = ON")
                cursor = conn.cursor()
                # Start transaction
                cursor.execute("BEGIN TRANSACTION")
                # Delete related records from Discipline
                cursor.execute("DELETE FROM Discipline WHERE teacher_id = ?", (user_id,))
                print("Discipline records deleted.")
                # Delete from Teacher table
                cursor.execute("DELETE FROM Teacher WHERE user_id = ?", (user_id,))
                print("Teacher record deleted.")
                # Delete from Users table
                cursor.execute("DELETE FROM Users WHERE id = ?", (user_id,))
                print("User record deleted.")
                conn.commit()
                QMessageBox.information(self, "Успех", "Преподаватель удален успешно.")
                self.update_lists() # Обновляем список преподавателей
        except sqlite3.IntegrityError as e:
            conn.rollback()
            QMessageBox.critical(self, "Ошибка", f"Не удалось удалить преподавателя: {e}")
        except sqlite3.Error as e:
            conn.rollback()
            QMessageBox.critical(self, "Ошибка", f"Ошибка базы данных: {e}")
        except Exception as e:
            conn.rollback()
            QMessageBox.critical(self, "Ошибка", f"Произошла ошибка: {e}")

    def create_add_group_tab(self):
        tab = QWidget()
        layout = QVBoxLayout()
        layout.setAlignment(Qt.AlignmentFlag.AlignCenter)

        group_box = QGroupBox("Добавить группу")
        group_box.setStyleSheet("QGroupBox { margin-top: 1ex; }"
                                "QGroupBox::title { subcontrol-origin: margin; left: 10px; }")

        form_layout = QFormLayout()
        form_layout.setLabelAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignTop)
        form_layout.setContentsMargins(20, 10, 20, 10)  # Adjust margins
        form_layout.setVerticalSpacing(20)  # Increase vertical spacing

        self.group_name_input = QLineEdit()
        self.group_name_input.setToolTip("Введите название группы")
        form_layout.addRow(QLabel("Название группы:"), self.group_name_input)

        self.speciality_name_input = QLineEdit()
        self.speciality_name_input.setToolTip("Введите специальность")
        form_layout.addRow(QLabel("Специальность:"), self.speciality_name_input)

        add_group_button = QPushButton("Добавить группу")
        add_group_button.clicked.connect(self.add_group)
        add_group_button.setToolTip("Нажмите, чтобы добавить новую группу")
        form_layout.addRow(add_group_button)

        group_box_layout = QVBoxLayout()
        group_box_layout.addLayout(form_layout)
        group_box.setLayout(group_box_layout)

        layout.addWidget(group_box)
        tab.setLayout(layout)
        return tab

    def add_group(self):
        group_name = self.group_name_input.text()
        speciality_name = self.speciality_name_input.text()

        # Проверка на пустые поля
        if not group_name or not speciality_name:
            QMessageBox.warning(self, "Ошибка", "Название группы и специальность не могут быть пустыми.")
            return

        try:
            with sqlite3.connect(self.db_path) as conn:
                cursor = conn.cursor()
                cursor.execute("INSERT INTO Group_ (group_name, speciality_name) VALUES (?, ?)",
                               (group_name, speciality_name))
                conn.commit()
                QMessageBox.information(self, "Успех", "Группа добавлена успешно.")
                self.group_name_input.clear()
                self.speciality_name_input.clear()
                self.update_lists() # Обновляем список групп для перемещения студентов
        except sqlite3.Error as e:
            QMessageBox.critical(self, "Ошибка", f"Ошибка базы данных: {e}")

    def create_delete_group_tab(self):
        tab = QWidget()
        layout = QVBoxLayout()

        group_box = QGroupBox("Удалить группу")
        group_box.setStyleSheet("QGroupBox { margin-top: 1ex; }"
                                "QGroupBox::title { subcontrol-origin: margin; left: 10px; }")

        form_layout = QFormLayout()
        form_layout.setLabelAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignTop)
        form_layout.setContentsMargins(20, 10, 20, 10)  # Adjust margins
        form_layout.setSpacing(30)  # Adjust spacing

        self.group_list_widget = QListWidget()
        self.load_groups_to_list(self.group_list_widget)
        self.group_list_widget.setToolTip("Выберите группу для удаления")
        form_layout.addRow(QLabel("Выберите группу:"), self.group_list_widget)

        delete_group_button = QPushButton("Удалить группу")
        delete_group_button.clicked.connect(self.delete_group_from_list)
        delete_group_button.setToolTip("Нажмите, чтобы удалить выбранную группу")
        form_layout.addRow(delete_group_button)

        group_box_layout = QVBoxLayout()
        group_box_layout.addLayout(form_layout)
        group_box.setLayout(group_box_layout)

        layout.addWidget(group_box)
        tab.setLayout(layout)
        return tab

    def load_groups_to_list(self, list_widget):
        list_widget.clear()
        try:
            with sqlite3.connect(self.db_path) as conn:
                cursor = conn.cursor()
                cursor.execute("SELECT group_id, group_name FROM Group_")
                self.groups_cache = cursor.fetchall()
                for group in self.groups_cache:
                    list_widget.addItem(group[1])
                    list_widget.item(list_widget.count() - 1).setData(Qt.ItemDataRole.UserRole, group[0])
        except sqlite3.Error as e:
            QMessageBox.critical(self, "Ошибка", f"Ошибка базы данных: {e}")

    def delete_group_from_list(self):
        selected_item = self.group_list_widget.currentItem()
        if not selected_item:
            QMessageBox.warning(self, "Ошибка", "Нет выбранной группы.")
            return
        group_id = selected_item.data(Qt.ItemDataRole.UserRole)
        self.delete_group(group_id)
        self.update_lists()

    def load_groups(self, combo_box):
        combo_box.clear()
        try:
            with sqlite3.connect(self.db_path) as conn:
                cursor = conn.cursor()
                cursor.execute("SELECT group_id, group_name FROM Group_")
                self.groups_cache = cursor.fetchall()
                for group in self.groups_cache:
                    combo_box.addItem(group[1], group[0])
        except sqlite3.Error as e:
            QMessageBox.critical(self, "Ошибка", f"Ошибка базы данных: {e}")

    def delete_group(self, group_id):
        if not self.show_question_message("Подтверждение", "Вы уверены, что хотите удалить эту группу?"):
            return

        try:
            with sqlite3.connect(self.db_path) as conn:
                conn.execute("PRAGMA foreign_keys = ON")
                cursor = conn.cursor()

                # Проверка наличия студентов в группе
                cursor.execute("SELECT COUNT(*) FROM Student WHERE group_id = ?", (group_id,))
                if cursor.fetchone()[0] > 0:
                    QMessageBox.warning(self, "Ошибка", "Нельзя удалить группу, в которой есть студенты.")
                    return

                # Start transaction
                cursor.execute("BEGIN TRANSACTION")
                # Delete related records from Discipline
                cursor.execute("DELETE FROM Discipline WHERE group_id = ?", (group_id,))
                print("Discipline records deleted.")
                # Delete from Group table
                cursor.execute("DELETE FROM Group_ WHERE group_id = ?", (group_id,))
                print("Group deleted.")
                conn.commit()
                QMessageBox.information(self, "Успех", "Группа удалена успешно.")
                self.update_lists()  # Обновляем список групп
        except sqlite3.IntegrityError as e:
            conn.rollback()
            QMessageBox.critical(self, "Ошибка", f"Не удалось удалить группу: {e}")
        except sqlite3.Error as e:
            conn.rollback()
            QMessageBox.critical(self, "Ошибка", f"Ошибка базы данных: {e}")
        except Exception as e:
            conn.rollback()
            QMessageBox.critical(self, "Ошибка", f"Произошла ошибка: {e}")

    def create_move_student_tab(self):
        tab = QWidget()
        layout = QVBoxLayout()
        layout.setAlignment(Qt.AlignmentFlag.AlignCenter)

        group_box = QGroupBox("Переместить студента")
        group_box.setStyleSheet("QGroupBox { margin-top: 1ex; }"
                                "QGroupBox::title { subcontrol-origin: margin; left: 10px; }")

        form_layout = QFormLayout()
        form_layout.setLabelAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignTop)
        form_layout.setContentsMargins(20, 10, 20, 10)  # Adjust margins
        form_layout.setVerticalSpacing(20)  # Increase vertical spacing

        self.move_student_list_widget = QListWidget()
        self.load_students_for_move(self.move_student_list_widget)
        self.move_student_list_widget.setToolTip("Выберите студента для перемещения")
        form_layout.addRow(QLabel("Выберите студента:"), self.move_student_list_widget)

        self.move_group_list_widget = QListWidget()
        self.load_groups_for_move(self.move_group_list_widget)
        self.move_group_list_widget.setToolTip("Выберите новую группу")
        form_layout.addRow(QLabel("Выберите новую группу:"), self.move_group_list_widget)

        move_student_button = QPushButton("Переместить студента")
        move_student_button.clicked.connect(self.move_student_from_list)
        move_student_button.setToolTip("Нажмите, чтобы переместить выбранного студента в новую группу")
        form_layout.addRow(move_student_button)

        group_box_layout = QVBoxLayout()
        group_box_layout.addLayout(form_layout)
        group_box.setLayout(group_box_layout)

        layout.addWidget(group_box)
        tab.setLayout(layout)
        return tab

    def load_students_for_move(self, widget):
        if isinstance(widget, QListWidget):
            widget.clear()
            for student in self.students_cache:
                display = f"{student[0]} - {student[1]} {student[2]}"
                widget.addItem(display)
                widget.item(widget.count() - 1).setData(Qt.ItemDataRole.UserRole, student[3])
        elif isinstance(widget, QComboBox):
            widget.clear()
            for student in self.students_cache:
                display = f"{student[0]} - {student[1]} {student[2]}"
                widget.addItem(display)

    def load_groups_for_move(self, list_widget):
        list_widget.clear()
        for group in self.groups_cache:
            list_widget.addItem(group[1])
            list_widget.item(list_widget.count() - 1).setData(Qt.ItemDataRole.UserRole, group[0])

    def move_student_from_list(self):
        selected_student_item = self.move_student_list_widget.currentItem()
        selected_group_item = self.move_group_list_widget.currentItem()
        if not selected_student_item or not selected_group_item:
            QMessageBox.warning(self, "Ошибка", "Нет выбранного студента или группы.")
            return
        student_id = selected_student_item.data(Qt.ItemDataRole.UserRole)
        group_id = selected_group_item.data(Qt.ItemDataRole.UserRole)

        # Проверка, находится ли студент уже в выбранной группе
        try:
            with sqlite3.connect(self.db_path) as conn:
                cursor = conn.cursor()
                cursor.execute("SELECT group_id FROM Student WHERE user_id = ?", (student_id,))
                current_group_id = cursor.fetchone()[0]
                if current_group_id == group_id:
                    QMessageBox.warning(self, "Ошибка", "Студент уже находится в этой группе.")
                    return
        except sqlite3.Error as e:
            QMessageBox.critical(self, "Ошибка", f"Ошибка базы данных: {e}")
            return

        if not self.show_question_message("Подтверждение", "Вы уверены, что хотите переместить этого студента?"):
            return
        try:
            with sqlite3.connect(self.db_path) as conn:
                cursor = conn.cursor()
                cursor.execute("UPDATE Student SET group_id = ? WHERE user_id = ?", (group_id, student_id))
                conn.commit()
                QMessageBox.information(self, "Успех", "Студент перемещен успешно.")
                self.load_students_for_move(self.move_student_list_widget)
                self.update_lists()
        except sqlite3.Error as e:
            QMessageBox.critical(self, "Ошибка", f"Ошибка базы данных: {e}")

    def move_student(self):
        selected_student_index = self.move_student_select_combo.currentIndex()
        selected_group_index = self.move_group_select_combo.currentIndex()
        if selected_student_index < 0 or selected_group_index < 0:
            QMessageBox.warning(self, "Ошибка", "Нет выбранного студента или группы.")
            return
        student_id = self.move_student_select_combo.itemData(selected_student_index)
        group_id = self.move_group_select_combo.itemData(selected_group_index)

        # Проверка, находится ли студент уже в выбранной группе
        try:
            with sqlite3.connect(self.db_path) as conn:
                cursor = conn.cursor()
                cursor.execute("SELECT group_id FROM Student WHERE user_id = ?", (student_id,))
                current_group_id = cursor.fetchone()[0]
                if current_group_id == group_id:
                    QMessageBox.warning(self, "Ошибка", "Студент уже находится в этой группе.")
                    return
        except sqlite3.Error as e:
            QMessageBox.critical(self, "Ошибка", f"Ошибка базы данных: {e}")
            return

        if not self.show_question_message("Подтверждение", "Вы уверены, что хотите переместить этого студента?"):
            return
        try:
            with sqlite3.connect(self.db_path) as conn:
                cursor = conn.cursor()
                cursor.execute("UPDATE Student SET group_id = ? WHERE user_id = ?", (group_id, student_id))
                conn.commit()
                QMessageBox.information(self, "Успех", "Студент перемещен успешно.")
                self.load_students_for_move(self.move_student_select_combo)
                self.update_lists()
        except sqlite3.Error as e:
            QMessageBox.critical(self, "Ошибка", f"Ошибка базы данных: {e}")

    def create_edit_student_tab(self):
        tab = QWidget()
        layout = QVBoxLayout()

        group_box = QGroupBox("Редактировать студента")
        group_box.setStyleSheet("QGroupBox { margin-top: 1ex; }"
                                "QGroupBox::title { subcontrol-origin: margin; left: 10px; }")

        form_layout = QFormLayout()
        form_layout.setLabelAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignTop)
        form_layout.setContentsMargins(20, 10, 20, 10)  # Adjust margins
        form_layout.setVerticalSpacing(20)  # Increase vertical spacing

        self.edit_student_select_combo = QComboBox()
        self.load_students(self.edit_student_select_combo)
        self.edit_student_select_combo.currentIndexChanged.connect(self.fill_student_fields)
        form_layout.addRow(QLabel("Выберите студента:"), self.edit_student_select_combo)

        self.edit_student_fname_input = QLineEdit()
        form_layout.addRow(QLabel("Имя:"), self.edit_student_fname_input)

        self.edit_student_lname_input = QLineEdit()
        form_layout.addRow(QLabel("Фамилия:"), self.edit_student_lname_input)

        self.edit_student_patronymic_input = QLineEdit()
        form_layout.addRow(QLabel("Отчество:"), self.edit_student_patronymic_input)

        self.edit_student_username_input = QLineEdit()
        form_layout.addRow(QLabel("Логин:"), self.edit_student_username_input)

        self.edit_student_password_input = QLineEdit()
        self.edit_student_password_input.setEchoMode(QLineEdit.EchoMode.Normal)  # Display plain text
        form_layout.addRow(QLabel("Пароль:"), self.edit_student_password_input)

        generate_password_button = QPushButton("Сгенерировать пароль")
        generate_password_button.clicked.connect(lambda: self.generate_password(self.edit_student_password_input))
        form_layout.addRow(generate_password_button)

        button_layout = QHBoxLayout()
        edit_student_button = QPushButton("Сохранить изменения")
        edit_student_button.clicked.connect(self.edit_student)
        button_layout.addWidget(edit_student_button)

        clear_fields_button = QPushButton("Очистить поля")
        clear_fields_button.clicked.connect(self.clear_edit_student_fields)
        button_layout.addWidget(clear_fields_button)

        form_layout.addRow(button_layout)

        group_box_layout = QVBoxLayout()
        group_box_layout.addLayout(form_layout)
        group_box.setLayout(group_box_layout)

        layout.addWidget(group_box)
        tab.setLayout(layout)
        return tab

    def clear_edit_student_fields(self):
        self.edit_student_fname_input.clear()
        self.edit_student_lname_input.clear()
        self.edit_student_patronymic_input.clear()
        self.edit_student_username_input.clear()
        self.edit_student_password_input.clear()

    def fill_student_fields(self):
        selected_index = self.edit_student_select_combo.currentIndex()
        if selected_index < 0:
            return
        user_id = self.edit_student_select_combo.itemData(selected_index)
        try:
            with sqlite3.connect(self.db_path) as conn:
                cursor = conn.cursor()
                cursor.execute("""
                    SELECT Users.username, Student.first_name, Student.last_name, Student.patronymic, Users.password
                    FROM Users
                    JOIN Student ON Users.id = Student.user_id
                    WHERE Users.id = ?
                """, (user_id,))
                student = cursor.fetchone()
                if student:
                    self.edit_student_username_input.setText(student[0])
                    self.edit_student_fname_input.setText(student[1])
                    self.edit_student_lname_input.setText(student[2])
                    self.edit_student_patronymic_input.setText(student[3])
                    self.edit_student_password_input.setText(student[4].decode('utf-8'))  # Преобразуем байты в строку
        except sqlite3.Error as e:
            QMessageBox.critical(self, "Ошибка", f"Ошибка базы данных: {e}")

    def create_edit_teacher_tab(self):
        tab = QWidget()
        layout = QVBoxLayout()

        group_box = QGroupBox("Редактировать преподавателя")
        group_box.setStyleSheet("QGroupBox { margin-top: 1ex; }"
                                "QGroupBox::title { subcontrol-origin: margin; left: 10px; }")

        form_layout = QFormLayout()
        form_layout.setLabelAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignTop)
        form_layout.setContentsMargins(20, 10, 20, 10)  # Adjust margins
        form_layout.setVerticalSpacing(20)  # Increase vertical spacing

        self.edit_teacher_select_combo = QComboBox()
        self.load_teachers(self.edit_teacher_select_combo)
        self.edit_teacher_select_combo.currentIndexChanged.connect(self.fill_teacher_fields)
        form_layout.addRow(QLabel("Выберите преподавателя:"), self.edit_teacher_select_combo)

        self.edit_teacher_fname_input = QLineEdit()
        form_layout.addRow(QLabel("Имя:"), self.edit_teacher_fname_input)

        self.edit_teacher_lname_input = QLineEdit()
        form_layout.addRow(QLabel("Фамилия:"), self.edit_teacher_lname_input)

        self.edit_teacher_patronymic_input = QLineEdit()
        form_layout.addRow(QLabel("Отчество:"), self.edit_teacher_patronymic_input)

        self.edit_teacher_username_input = QLineEdit()
        form_layout.addRow(QLabel("Логин:"), self.edit_teacher_username_input)

        self.edit_teacher_password_input = QLineEdit()
        self.edit_teacher_password_input.setEchoMode(QLineEdit.EchoMode.Normal)  # Display plain text
        form_layout.addRow(QLabel("Пароль:"), self.edit_teacher_password_input)

        generate_password_button = QPushButton("Сгенерировать пароль")
        generate_password_button.clicked.connect(lambda: self.generate_password(self.edit_teacher_password_input))
        form_layout.addRow(generate_password_button)

        button_layout = QHBoxLayout()
        edit_teacher_button = QPushButton("Сохранить изменения")
        edit_teacher_button.clicked.connect(self.edit_teacher)
        button_layout.addWidget(edit_teacher_button)

        clear_fields_button = QPushButton("Очистить поля")
        clear_fields_button.clicked.connect(self.clear_edit_teacher_fields)
        button_layout.addWidget(clear_fields_button)

        form_layout.addRow(button_layout)

        group_box_layout = QVBoxLayout()
        group_box_layout.addLayout(form_layout)
        group_box.setLayout(group_box_layout)

        layout.addWidget(group_box)
        tab.setLayout(layout)
        return tab

    def clear_edit_teacher_fields(self):
        self.edit_teacher_fname_input.clear()
        self.edit_teacher_lname_input.clear()
        self.edit_teacher_patronymic_input.clear()
        self.edit_teacher_username_input.clear()
        self.edit_teacher_password_input.clear()

    def fill_teacher_fields(self):
        selected_index = self.edit_teacher_select_combo.currentIndex()
        if selected_index < 0:
            return
        user_id = self.edit_teacher_select_combo.itemData(selected_index)
        try:
            with sqlite3.connect(self.db_path) as conn:
                cursor = conn.cursor()
                cursor.execute("""
                    SELECT Users.username, Teacher.first_name, Teacher.last_name, Teacher.patronymic, Users.password
                    FROM Users
                    JOIN Teacher ON Users.id = Teacher.user_id
                    WHERE Users.id = ?
                """, (user_id,))
                teacher = cursor.fetchone()
                if teacher:
                    self.edit_teacher_username_input.setText(teacher[0])
                    self.edit_teacher_fname_input.setText(teacher[1])
                    self.edit_teacher_lname_input.setText(teacher[2])
                    self.edit_teacher_patronymic_input.setText(teacher[3])
                    self.edit_teacher_password_input.setText(teacher[4].decode('utf-8'))  # Преобразуем байты в строку
        except sqlite3.Error as e:
            QMessageBox.critical(self, "Ошибка", f"Ошибка базы данных: {e}")

    def edit_student(self):
        selected_index = self.edit_student_select_combo.currentIndex()
        if selected_index < 0:
            QMessageBox.warning(self, "Ошибка", "Нет выбранного студента.")
            return
        user_id = self.edit_student_select_combo.itemData(selected_index)
        fname = self.edit_student_fname_input.text()
        lname = self.edit_student_lname_input.text()
        patronymic = self.edit_student_patronymic_input.text()
        login = self.edit_student_username_input.text()
        password = self.edit_student_password_input.text()

        # Проверка на одинаковые логины
        with sqlite3.connect(self.db_path) as conn:
            cursor = conn.cursor()
            cursor.execute("SELECT id FROM Users WHERE username = ? AND id != ?", (login, user_id))
            if cursor.fetchone():
                QMessageBox.warning(self, "Ошибка", "Пользователь с таким логином уже существует.")
                return

        # Проверка на пустые поля
        if not fname or not lname or not patronymic or not login:
            QMessageBox.warning(self, "Ошибка", "Все поля должны быть заполнены.")
            return

        # Проверка на текстовые поля
        if not fname.isalpha() or not lname.isalpha() or not patronymic.isalpha():
            QMessageBox.warning(self, "Ошибка", "Поля 'Имя', 'Фамилия' и 'Отчество' должны содержать только текст.")
            return

        try:
            with sqlite3.connect(self.db_path) as conn:
                cursor = conn.cursor()
                # Update Users table
                if password:
                    hashed_password = bcrypt.hashpw(password.encode(), bcrypt.gensalt())
                    cursor.execute("UPDATE Users SET username = ?, password = ? WHERE id = ?",
                                   (login, hashed_password, user_id))
                else:
                    cursor.execute("UPDATE Users SET username = ? WHERE id = ?", (login, user_id))

                # Update Student table
                cursor.execute("UPDATE Student SET first_name = ?, last_name = ?, patronymic = ? WHERE user_id = ?",
                               (fname, lname, patronymic, user_id))
                conn.commit()
                QMessageBox.information(self, "Успех", "Данные студента обновлены успешно.")
                self.load_students(self.edit_student_select_combo)
                self.update_lists()
        except sqlite3.Error as e:
            QMessageBox.critical(self, "Ошибка", f"Ошибка базы данных: {e}")

    def edit_teacher(self):
        selected_index = self.edit_teacher_select_combo.currentIndex()
        if selected_index < 0:
            QMessageBox.warning(self, "Ошибка", "Нет выбранного преподавателя.")
            return
        user_id = self.edit_teacher_select_combo.itemData(selected_index)
        fname = self.edit_teacher_fname_input.text()
        lname = self.edit_teacher_lname_input.text()
        patronymic = self.edit_teacher_patronymic_input.text()
        login = self.edit_teacher_username_input.text()
        password = self.edit_teacher_password_input.text()

        # Проверка на одинаковые логины
        with sqlite3.connect(self.db_path) as conn:
            cursor = conn.cursor()
            cursor.execute("SELECT id FROM Users WHERE username = ? AND id != ?", (login, user_id))
            if cursor.fetchone():
                QMessageBox.warning(self, "Ошибка", "Пользователь с таким логином уже существует.")
                return

        # Проверка на пустые поля
        if not fname or not lname or not patronymic or not login:
            QMessageBox.warning(self, "Ошибка", "Все поля должны быть заполнены.")
            return

        # Проверка на текстовые поля
        if not fname.isalpha() or not lname.isalpha() or not patronymic.isalpha():
            QMessageBox.warning(self, "Ошибка", "Поля 'Имя', 'Фамилия' и 'Отчество' должны содержать только текст.")
            return

        try:
            with sqlite3.connect(self.db_path) as conn:
                cursor = conn.cursor()
                # Update Users table
                if password:
                    hashed_password = bcrypt.hashpw(password.encode(), bcrypt.gensalt())
                    cursor.execute("UPDATE Users SET username = ?, password = ? WHERE id = ?",
                                   (login, hashed_password, user_id))
                else:
                    cursor.execute("UPDATE Users SET username = ? WHERE id = ?", (login, user_id))

                # Update Teacher table
                cursor.execute("UPDATE Teacher SET first_name = ?, last_name = ?, patronymic = ? WHERE user_id = ?",
                               (fname, lname, patronymic, user_id))
                conn.commit()
                QMessageBox.information(self, "Успех", "Данные преподавателя обновлены успешно.")
                self.load_teachers(self.edit_teacher_select_combo)
                self.update_lists()
        except sqlite3.Error as e:
            QMessageBox.critical(self, "Ошибка", f"Ошибка базы данных: {e}")














class TeacherWindow(QMainWindow):
    def __init__(self, teacher_id, db_path):
        super().__init__()
        self.teacher_id = teacher_id
        self.db_path = db_path
        self.setWindowTitle("Окно преподавателя")
        self.central_widget = QWidget()
        self.setCentralWidget(self.central_widget)
        self.journal_widget = None
        self.discipline_widget = None
        self.current_discipline_id = None
        self.dates = []  # Инициализация атрибута dates

        self.conn = sqlite3.connect(self.db_path)
        self.cursor = self.conn.cursor()
        self.initUI()

    def closeEvent(self, event):
        self.conn.close()
        super().closeEvent(event)

    def initUI(self):
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        self.main_layout = QVBoxLayout(central_widget)
        self.main_layout.setAlignment(Qt.AlignmentFlag.AlignTop)
        self.main_layout.setContentsMargins(20, 20, 20, 20)

        # Получаем ФИО преподавателя из базы данных
        try:
            with sqlite3.connect(self.db_path) as conn:
                cursor = conn.cursor()
                cursor.execute("SELECT first_name, last_name, patronymic FROM Teacher WHERE user_id = ?",
                               (self.teacher_id,))
                result = cursor.fetchone()
                if result:
                    first_name, last_name, patronymic = result
                    teacher_name = f"{last_name} {first_name} {patronymic}"
                    teacher_label = QLabel(f"Преподаватель: {teacher_name}")
                    teacher_label.setFont(QFont("Arial", 24, QFont.Weight.Bold))
                    teacher_label.setStyleSheet("color: black;")
                    teacher_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
                    self.main_layout.addWidget(teacher_label)
                else:
                    QMessageBox.warning(self, "Ошибка", "Преподаватель не найден.")
        except sqlite3.Error as e:
            QMessageBox.critical(self, "Ошибка", f"Ошибка базы данных: {e}")

        group_layout = QHBoxLayout()
        group_layout.setSpacing(0)
        group_layout.setContentsMargins(0, 0, 0, 0)
        group_label = QLabel("Группа:")
        group_label.setFont(QFont("Arial", 18))
        group_label.setStyleSheet("color: black;")
        group_label.setFixedWidth(72)
        group_layout.addWidget(group_label)

        # Получаем номер группы, которую ведет преподаватель
        try:
            with sqlite3.connect(self.db_path) as conn:
                cursor = conn.cursor()
                cursor.execute("""
                            SELECT G.group_name 
                            FROM Teacher T 
                            JOIN Discipline D ON T.user_id = D.teacher_id 
                            JOIN Group_ G ON D.group_id = G.group_id 
                            WHERE T.user_id = ?
                        """, (self.teacher_id,))
                result = cursor.fetchone()
                if result:
                    group_name = result[0]
                    group_button = QPushButton(group_name)
                    group_button.setFont(QFont("Arial", 18))
                    group_button.setStyleSheet("""
                                QPushButton {
                                    background-color: {lightblue};
                                    border: 1px solid {darkblue};
                                    border-radius: 4px;
                                    padding: 2px 4px;
                                    min-height: 20px;
                                    color: white;
                                    font-weight: bold;
                                }

                                QPushButton:hover {
                                    background-color: {darkblue};
                                    border: 1px solid {lightblue};
                                }

                                QPushButton:pressed {
                                    background-color: {darkblue};
                                }
                            """)
                    group_button.setFixedWidth(80)
                    group_button.clicked.connect(self.show_disciplines)
                    group_layout.addWidget(group_button)
                else:
                    QMessageBox.warning(self, "Ошибка", "Группа не найдена.")
        except sqlite3.Error as e:
            QMessageBox.critical(self, "Ошибка", f"Ошибка базы данных: {e}")

        top_layout = QHBoxLayout()
        top_layout.addStretch()
        top_layout.addLayout(group_layout)
        top_layout.addStretch()

        self.main_layout.addLayout(top_layout)

        self.discipline_layout = QVBoxLayout()
        self.main_layout.addLayout(self.discipline_layout)

        self.journal_area = QWidget()
        self.journal_area_layout = QVBoxLayout()
        self.journal_area_layout.setAlignment(Qt.AlignmentFlag.AlignTop)
        self.journal_area.setLayout(self.journal_area_layout)
        self.main_layout.addWidget(self.journal_area)
        self.journal_area.hide()

        self.setMinimumSize(1150, 750)

    def show_disciplines(self):
        self.journal_area.hide()
        if self.journal_widget is not None:
            self.journal_area_layout.removeWidget(self.journal_widget)
            self.journal_widget.deleteLater()
            self.journal_widget = None

        for i in reversed(range(self.discipline_layout.count())):
            widget = self.discipline_layout.itemAt(i).widget()
            if widget is not None:
                self.discipline_layout.removeWidget(widget)
                widget.deleteLater()
        try:
            with sqlite3.connect(self.db_path) as conn:
                cursor = conn.cursor()
                cursor.execute("""
                        SELECT D.discipline_name 
                        FROM Teacher T 
                        JOIN Discipline D ON T.user_id = D.teacher_id 
                        WHERE T.user_id = ?
                    """, (self.teacher_id,))
                disciplines = cursor.fetchall()
                for discipline in disciplines:
                    discipline_button = QPushButton(discipline[0])
                    discipline_button.setFont(QFont("Arial", 18))
                    discipline_button.setStyleSheet("""
                            QPushButton {
                                background-color: {lightblue};
                                border: 1px solid {darkblue};
                                border-radius: 4px;
                                padding: 2px 4px;
                                min-height: 20px;
                                color: white;
                                font-weight: bold;
                            }

                            QPushButton:hover {
                                background-color: {darkblue};
                                border: 1px solid {lightblue};
                            }

                            QPushButton:pressed {
                                background-color: {darkblue};
                            }
                        """)
                    discipline_button.clicked.connect(
                        lambda checked, discipline_name=discipline[0]: self.show_journal(discipline_name))
                    self.discipline_layout.addWidget(discipline_button)

        except sqlite3.Error as e:
            QMessageBox.critical(self, "Ошибка", f"Ошибка базы данных: {e}")

    def save_journal(self):
        try:
            self.conn.commit()  # Сохраняем изменения
            QMessageBox.information(self, "Успех", "Журнал успешно сохранен.")
            self.refresh_journal()  # Обновляем журнал
        except sqlite3.Error as e:
            QMessageBox.critical(self, "Ошибка", f"Ошибка при сохранении журнала: {e}")

    def refresh_journal(self):
        if hasattr(self, 'current_discipline_name'):  # Проверяем, установлено ли имя дисциплины
            self.show_journal(self.current_discipline_name)

    def show_journal(self, discipline_name):
        self.journal_area.show()
        if self.journal_widget is not None:
            self.journal_area_layout.removeWidget(self.journal_widget)
            self.journal_widget.deleteLater()

        self.journal_widget = QWidget()
        self.journal_widget_layout = QVBoxLayout()
        self.journal_widget.setLayout(self.journal_widget_layout)

        try:
            with sqlite3.connect(self.db_path) as conn:
                cursor = self.conn.cursor()

                # Получаем ID дисциплины
                cursor.execute("SELECT discipline_id FROM Discipline WHERE discipline_name = ?", (discipline_name,))
                discipline_id_result = cursor.fetchone()
                if not discipline_id_result:
                    QMessageBox.warning(self, "Ошибка", f"Дисциплина '{discipline_name}' не найдена.")
                    return
                discipline_id = discipline_id_result[0]
                self.current_discipline_id = discipline_id
                self.current_discipline_name = discipline_name  # Сохраняем имя дисциплины

                # Получаем уникальные даты посещений для данной дисциплины
                cursor.execute("""
                    SELECT DISTINCT attendance_date 
                    FROM Attendance
                    WHERE discipline_id = ?
                    ORDER BY attendance_date
                """, (discipline_id,))
                self.dates = [row[0] for row in cursor.fetchall()]

                # Получаем список студентов для этой дисциплины
                cursor.execute("""
                    SELECT S.user_id, S.last_name, S.first_name, S.patronymic
                    FROM Student S
                    JOIN Group_ G ON S.group_id = G.group_id
                    JOIN Discipline D ON G.group_id = D.group_id
                    WHERE D.discipline_id = ?
                    ORDER BY S.last_name, S.first_name, S.patronymic
                """, (discipline_id,))
                students = cursor.fetchall()

                # Создаем таблицу
                label = QLabel(f"Журнал по {discipline_name}:")
                label.setFont(QFont("Arial", 18))
                label.setAlignment(Qt.AlignmentFlag.AlignCenter)
                self.journal_widget_layout.addWidget(label)

                journal_table = QTableWidget()
                journal_table.setColumnCount(len(self.dates) + 2)
                journal_table.setRowCount(len(students))

                # Добавляем заголовки
                header_labels = ["Студент"] + self.dates + ["Средняя оценка"]
                journal_table.setHorizontalHeaderLabels(header_labels)

                # Заполняем таблицу данными
                for row, student in enumerate(students):
                    student_id, last_name, first_name, patronymic = student
                    student_name = f"{last_name} {first_name} {patronymic}"
                    item = QTableWidgetItem(student_name)
                    item.setFlags(item.flags() & ~Qt.ItemFlag.ItemIsEditable)
                    journal_table.setItem(row, 0, item)

                    # Заполняем оценки и статусы посещаемости для каждого студента по каждой дате
                    for col, date in enumerate(self.dates):
                        # Получаем информацию о посещаемости
                        cursor.execute("""
                            SELECT is_present
                            FROM Attendance
                            WHERE student_id = ? AND discipline_id = ? AND attendance_date = ?
                        """, (student_id, discipline_id, date))
                        attendance_result = cursor.fetchone()

                        cursor.execute("""
                            SELECT grade_value, grade_type
                            FROM Grade
                            WHERE student_id = ? AND discipline_id = ? AND grade_date = ?
                        """, (student_id, discipline_id, date))
                        grade_result = cursor.fetchone()

                        if attendance_result and attendance_result[0] == 0:  # Проверяем is_present == 0
                            item_text = "Н"
                        elif grade_result:
                            grade_value, grade_type = grade_result
                            item_text = f"{grade_value} ({grade_type})" if grade_value is not None else ""
                        else:
                            item_text = ""  # Оставляем пустую строку, если нет ни оценки, ни "Н"

                        item = QTableWidgetItem(item_text)
                        item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                        journal_table.setItem(row, col + 1, item)

                        # Разрешаем редактирование ячеек
                        item.setFlags(item.flags() | Qt.ItemFlag.ItemIsEditable)

                    # Расчет и добавление средней оценки с учетом "удвоенных" оценок за контрольные
                    cursor.execute("""
                        SELECT grade_value, grade_type
                        FROM Grade
                        WHERE student_id = ? AND discipline_id = ?
                    """, (student_id, discipline_id))
                    grades_results = cursor.fetchall()
                    grades = []
                    for grade_value, grade_type in grades_results:
                        if grade_value is not None:
                            grades.append(grade_value)
                            if grade_type == "Контрольная работа":
                                grades.append(grade_value)  # Добавляем оценку еще раз, если это контрольная

                    average_grade = sum(grades) / len(grades) if grades else 0

                    average_grade_item = QTableWidgetItem(f"{average_grade:.2f}")
                    average_grade_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                    average_grade_item.setFlags(average_grade_item.flags() & ~Qt.ItemFlag.ItemIsEditable)
                    journal_table.setItem(row, len(self.dates) + 1, average_grade_item)

                journal_table.itemChanged.connect(self.validate_item)
                journal_table.itemChanged.connect(self.on_item_changed)

                journal_table.resizeColumnsToContents()
                self.journal_widget_layout.addWidget(journal_table)
                self.journal_area_layout.addWidget(self.journal_widget)


                # Создаем горизонтальный макет для кнопок
                button_layout = QHBoxLayout()
                button_layout.setSpacing(10)  # Устанавливаем отступ между кнопками

                save_button = QPushButton("Сохранить журнал")
                save_button.clicked.connect(self.save_journal)
                save_button.setFixedSize(500, 50)  # Устанавливаем фиксированный размер кнопки
                button_layout.addWidget(save_button)

                export_button = QPushButton("Экспорт в Excel")
                export_button.clicked.connect(self.export_to_excel)
                export_button.setFixedSize(500, 50)  # Устанавливаем фиксированный размер кнопки
                button_layout.addWidget(export_button)

                self.journal_widget_layout.addLayout(button_layout)

        except sqlite3.Error as e:
            QMessageBox.critical(self, "Ошибка", f"Ошибка базы данных: {e}")

    def validate_item(self, item):
        # Разрешаем изменение даже если ячейка остаётся пустой
        if item.column() > 0 and item.column() < len(self.dates) + 1:
            text = item.text()
            if not text.strip():
                return

            if not re.match("^(2|3|4|5|Н)$", text):
                # Если не оценка и не "Н", то очищаем ячейку и показываем предупреждение
                QMessageBox.warning(self, "Ошибка", "В ячейке может быть только оценка от 2 до 5 или буква 'Н'.")
                item.setText("")

    def on_item_changed(self, item):
        """
        Handles cell changes in the journal table.
        Saves grade or attendance status to the database.
        """
        try:
            if not self.conn:
                QMessageBox.critical(self, "Ошибка", "Соединение с базой данных не установлено.")
                return
            row = item.row()
            col = item.column()

            if col == 0 or col == item.tableWidget().columnCount() - 1:
                return

            student_name = item.tableWidget().item(row, 0).text()

            self.cursor.execute("""
                SELECT user_id FROM Student
                WHERE last_name || ' ' || first_name || ' ' || patronymic = ?
            """, (student_name,))
            result = self.cursor.fetchone()
            if not result:
                QMessageBox.warning(self, "Ошибка", "Не удалось найти студента в базе данных.")
                return
            student_id = result[0]

            date = item.tableWidget().horizontalHeaderItem(col).text()
            new_value = item.text().strip()

            if new_value == "Н":
                # Обработка посещаемости
                self.cursor.execute("""
                    SELECT is_present
                    FROM Attendance
                    WHERE student_id = ? AND discipline_id = ? AND attendance_date = ?
                """, (student_id, self.current_discipline_id, date))
                attendance_exists = self.cursor.fetchone()

                if attendance_exists:  # Если запись существует, обновляем её
                    self.cursor.execute("""
                        UPDATE Attendance
                        SET is_present = 0
                        WHERE student_id = ? AND discipline_id = ? AND attendance_date = ?
                    """, (student_id, self.current_discipline_id, date))
                else:  # Если записи нет, создаём новую
                    self.cursor.execute("""
                        INSERT INTO Attendance (student_id, discipline_id, attendance_date, is_present)
                        VALUES (?, ?, ?, 0)
                    """, (student_id, self.current_discipline_id, date))

                # Удаляем оценки, если были
                self.cursor.execute("""
                    DELETE FROM Grade
                    WHERE student_id = ? AND discipline_id = ? AND grade_date = ?
                """, (student_id, self.current_discipline_id, date))

            elif new_value.isdigit() and 2 <= int(new_value) <= 5:
                # Обработка оценки
                grade_types = ["Домашняя работа", "Контрольная работа", "Ответ на паре", "Практическая работа",
                               "Лабораторная работа"]
                grade_type, ok = QInputDialog.getItem(self, "Тип оценки", "Выберите тип оценки:", grade_types, 0, False)

                if ok:
                    # Проверяем, существует ли уже оценка для этого студента на эту дату
                    self.cursor.execute("""
                        SELECT grade_value
                        FROM Grade
                        WHERE student_id = ? AND discipline_id = ? AND grade_date = ?
                    """, (student_id, self.current_discipline_id, date))
                    grade_exists = self.cursor.fetchone()

                    if grade_exists:  # Если запись существует, обновляем её
                        self.cursor.execute("""
                            UPDATE Grade
                            SET grade_value = ?, grade_type = ?
                            WHERE student_id = ? AND discipline_id = ? AND grade_date = ?
                        """, (int(new_value), grade_type, student_id, self.current_discipline_id, date))
                    else:  # Если записи нет, создаём новую
                        self.cursor.execute("""
                            INSERT INTO Grade (student_id, discipline_id, grade_date, grade_value, grade_type)
                            VALUES (?, ?, ?, ?, ?)
                        """, (student_id, self.current_discipline_id, date, int(new_value), grade_type))

                    # Обновляем посещаемость
                    self.cursor.execute("""
                        INSERT OR REPLACE INTO Attendance (student_id, discipline_id, attendance_date, is_present)
                        VALUES (?, ?, ?, 1)
                    """, (student_id, self.current_discipline_id, date))

            else:  # Обработка пустой строки
                # Удаляем оценку, если она была
                self.cursor.execute("""
                    DELETE FROM Grade
                    WHERE student_id = ? AND discipline_id = ? AND grade_date = ?
                """, (student_id, self.current_discipline_id, date))

                # Удаляем посещаемость
                self.cursor.execute("""
                    UPDATE Attendance
                    SET is_present = NULL
                    WHERE student_id = ? AND discipline_id = ? AND attendance_date = ?
                """, (student_id, self.current_discipline_id, date))

            self.conn.commit()  # Сохраняем изменения
            self.refresh_journal()  # Обновляем журнал

        except sqlite3.Error as e:
            QMessageBox.critical(self, "Ошибка", f"Ошибка базы данных: {e}")

    def export_to_excel(self):
        """
        Экспортирует текущий журнал в Excel-файл.
        """
        if not self.journal_widget:
            QMessageBox.warning(self, "Ошибка", "Журнал не отображается.")
            return

        # Получаем таблицу из журнала
        journal_table = self.journal_widget_layout.itemAt(1).widget()
        if not isinstance(journal_table, QTableWidget):
            QMessageBox.warning(self, "Ошибка", "Не удалось найти таблицу журнала.")
            return

        # Создаем новый Excel-файл
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Журнал"

        # Заполняем заголовки
        headers = [journal_table.horizontalHeaderItem(col).text() for col in range(journal_table.columnCount())]
        for col_num, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col_num, value=header)
            cell.alignment = Alignment(horizontal='center', wrap_text=True)

        # Заполняем данные
        for row_num in range(journal_table.rowCount()):
            for col_num in range(journal_table.columnCount()):
                item = journal_table.item(row_num, col_num)
                if item:
                    cell = sheet.cell(row=row_num + 2, column=col_num + 1, value=item.text())
                    cell.alignment = Alignment(horizontal='center', wrap_text=True)

        # Устанавливаем автоширину для столбцов
        for col_num in range(1, journal_table.columnCount() + 1):
            column_letter = openpyxl.utils.get_column_letter(col_num)
            sheet.column_dimensions[column_letter].auto_size = True

        # Запрашиваем путь для сохранения файла
        file_path, _ = QFileDialog.getSaveFileName(self, "Сохранить журнал", "", "Excel files (*.xlsx);;All files (*)")
        if file_path:
            workbook.save(file_path)
            QMessageBox.information(self, "Успех", f"Журнал успешно экспортирован в файл: {file_path}")



















class MainWindow(QMainWindow):
    def __init__(self, db_path="College.db"):
        super().__init__()
        self.db_path = db_path
        self.setWindowTitle("Авторизация")
        self.error_message = QErrorMessage(self)
        self.central_widget = QWidget()
        self.setCentralWidget(self.central_widget)
        self.initUI()

        # Apply the stylesheet here, before showing the window
        app.setStyleSheet(app_style)

        self.setMinimumSize(QSize(400, 300))

    def initUI(self):
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        main_layout = QVBoxLayout(central_widget)
        main_layout.setAlignment(Qt.AlignmentFlag.AlignCenter)
        main_layout.setContentsMargins(20, 20, 20, 20)

        auth_label = QLabel("Авторизация")
        auth_label.setFont(QFont("Arial", 18, QFont.Weight.Bold))
        auth_label.setStyleSheet("color: black;")  # Устанавливаем цвет текста на черный
        auth_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        main_layout.addWidget(auth_label)

        frame = QFrame()
        frame.setFrameShape(QFrame.Shape.StyledPanel)
        frame_layout = QVBoxLayout(frame)
        frame_layout.setContentsMargins(20, 15, 20, 15)

        form_layout = QFormLayout()
        form_layout.setLabelAlignment(Qt.AlignmentFlag.AlignRight)
        form_layout.setHorizontalSpacing(10)
        form_layout.setVerticalSpacing(15)

        # Определяем поля ввода
        self.login_edit = QLineEdit()
        self.password_edit = QLineEdit()
        self.password_edit.setEchoMode(QLineEdit.EchoMode.Password)

        # Фиксируем ширину полей ввода
        self.login_edit.setFixedWidth(200)
        self.password_edit.setFixedWidth(200)

        # Используем QHBoxLayout для каждой пары "метка-поле ввода"
        login_layout = QHBoxLayout()
        login_label = QLabel("Логин:")
        login_label.setFont(QFont("Arial", 12, QFont.Weight.Bold))
        login_label.setStyleSheet("color: black;")  # Устанавливаем цвет текста на черный
        login_layout.addWidget(login_label)
        login_layout.addSpacing(5)  # Уменьшаем отступ между меткой и полем ввода
        login_layout.addWidget(self.login_edit)
        login_layout.setAlignment(Qt.AlignmentFlag.AlignCenter)
        form_layout.addRow(login_layout)

        password_layout = QHBoxLayout()
        password_label = QLabel("Пароль:")
        password_label.setFont(QFont("Arial", 12, QFont.Weight.Bold))
        password_label.setStyleSheet("color: black;")  # Устанавливаем цвет текста на черный
        password_layout.addWidget(password_label)
        password_layout.addSpacing(5)  # Уменьшаем отступ между меткой и полем ввода
        password_layout.addWidget(self.password_edit)
        password_layout.setAlignment(Qt.AlignmentFlag.AlignCenter)
        form_layout.addRow(password_layout)

        frame_layout.addLayout(form_layout)
        main_layout.addWidget(frame)

        login_button = QPushButton("Войти")
        login_button.setFont(QFont("Arial", 12, QFont.Weight.Bold))
        login_button.setFixedWidth(150)
        login_button.clicked.connect(self.login)
        main_layout.addWidget(login_button, alignment=Qt.AlignmentFlag.AlignCenter)
    def login(self):
        username = self.login_edit.text()
        password = self.password_edit.text()

        if not username:
            self.error_message.showMessage("Поле 'Логин' не может быть пустым")
            return

        if not password:
            self.error_message.showMessage("Поле 'Пароль' не может быть пустым")
            return

        if username == "admin":
            hashed_password = b'$2b$12$Txzt6mH46xIHIv0uBbcRk.c38.zceHSzGigFBEe72q4zHKFBAf8jK'
            if bcrypt.checkpw(password.encode(), hashed_password):
                QMessageBox.information(self, "Успех", "Вход выполнен (администратор)")
                self.admin_window = AdminWindow(db_path=self.db_path)
                self.admin_window.show()
                self.close()
            else:
                QMessageBox.warning(self, "Ошибка", "Неверный пароль")
        else:
            try:
                with sqlite3.connect(self.db_path) as conn:
                    cursor = conn.cursor()
                    cursor.execute("SELECT password, role, id FROM Users WHERE lower(username) = ?",
                                   (username.lower(),))
                    result = cursor.fetchone()
                    if result:
                        stored_password, role, user_id = result
                        if bcrypt.checkpw(password.encode(), stored_password):
                            if role == 'teacher':
                                # Fetch teacher_id from Teacher table
                                cursor.execute("SELECT user_id FROM Teacher WHERE user_id = ?", (user_id,))
                                teacher_result = cursor.fetchone()
                                if teacher_result:
                                    teacher_id = teacher_result[0]
                                    QMessageBox.information(self, "Успех", "Вход выполнен (преподаватель)")
                                    self.teacher_window = TeacherWindow(teacher_id=teacher_id, db_path=self.db_path)
                                    self.teacher_window.show()
                                    self.close()
                                else:
                                    QMessageBox.warning(self, "Ошибка", "Преподаватель не найден.")
                            elif role == 'student':
                                # Fetch student_id from Student table
                                cursor.execute("SELECT user_id FROM Student WHERE user_id = ?", (user_id,))
                                student_result = cursor.fetchone()
                                if student_result:
                                    student_id = student_result[0]
                                    QMessageBox.information(self, "Успех", "Вход выполнен (студент)")
                                    self.student_window = StudentWindow(student_id=student_id, db_path=self.db_path)
                                    self.student_window.show()
                                    self.close()
                                else:
                                    QMessageBox.warning(self, "Ошибка", "Студент не найден.")
                            else:
                                QMessageBox.warning(self, "Ошибка", "Неизвестная роль пользователя.")
                        else:
                            QMessageBox.warning(self, "Ошибка", "Неверный пароль")
                    else:
                        QMessageBox.warning(self, "Ошибка", "Пользователь не найден")
            except sqlite3.Error as e:
                QMessageBox.critical(self, "Ошибка", f"Ошибка базы данных: {e}")



if __name__ == "__main__":
    app = QApplication(sys.argv)
    app.setStyleSheet(app_style)  # Применяем стиль к приложению
    window = MainWindow()
    window.show()
    sys.exit(app.exec())
